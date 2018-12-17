#!usr/bin/env python
import fnmatch
import os
import sys
import time

import openpyxl
import validators
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchFrameException
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

excel_file = None

# Determine if application is a script file or frozen exe
if getattr(sys, 'frozen', False):
    application_path = os.path.dirname(sys.executable)
elif __file__:
    application_path = os.path.dirname(__file__)

for file in os.listdir(application_path):
    if fnmatch.fnmatch(file, '*.xlsx'):
        excel_file = file
        print('Opening file: ' + '\'' + file.title() + '\'')
if excel_file is not None:
    wb = openpyxl.load_workbook(application_path + '/' + excel_file)
else:
    sys.exit('No excel file found')

# First sheet
ws = wb[wb.sheetnames[0]]

url = "https://property.phila.gov"

# Selenium Chrome Webdriver
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('windows-size=1200x600')
chrome_options.add_argument('--disable-gpu')

# dir_path = os.path.dirname(os.path.realpath(__file__))

driver = webdriver.Chrome(application_path + '/chromedriver', options=chrome_options)

column_with_addresses = "A"
column_owner = "H"
column_mailing = "G"
column_description = "I"

for row_number in range(2, ws.max_row + 1):
    try:
        address = ws[column_with_addresses + str(row_number)].value.strip()
    except Exception:
        exit('Failed to access cell at ' + column_with_addresses + str(row_number) + ', exiting...')

    print('Row number: ' + str(row_number))
    print('Checking address: ' + address)

    driver.get(url)

    input = driver.find_element_by_id('search-address')
    input.send_keys(address)

    input.send_keys(Keys.ENTER)
    time.sleep(2)
    owners_element = ''
    mailing_element = ''
    description_element = ''

    try:

        owners_element = driver.find_element_by_class_name("owner-name")
        mailing_element = driver.find_element_by_css_selector("div.bold.mailing")
        description_element = driver.find_element_by_xpath(
            '//*[@id="maincontent"]/div[3]/div[2]/div[2]/div[3]/div/strong')
    except Exception:
        print('Prolonging...')
        time.sleep(2)
        try:
            owners_element = driver.find_element_by_class_name("owner-name")
            mailing_element = driver.find_element_by_css_selector("div.bold.mailing")
            description_element = driver.find_element_by_xpath(
                '//*[@id="maincontent"]/div[3]/div[2]/div[2]/div[3]/div/strong')
        except Exception:
            print('Failed to find results on row number: ' + str(row_number) + ', skipping row...')
            continue

    print('Owner found: ' + owners_element.text)
    print('Description found: ' + description_element.text)
    print('Address found: ' + mailing_element.text + '\n')

    ws[column_owner + str(row_number)] = owners_element.text
    ws[column_mailing + str(row_number)] = mailing_element.text
    ws[column_description + str(row_number)] = description_element.text
    wb.save(application_path + '/' + excel_file.title())

driver.quit()
